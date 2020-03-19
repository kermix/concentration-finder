import json
from itertools import zip_longest

import numpy as np
import pandas as pd
from openpyxl.styles import Font, Alignment

from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

from openpyxl.worksheet.table import Table

from openpyxl.utils.cell import (
    column_index_from_string,
    coordinate_from_string,
    get_column_letter
)

from openpyxl import drawing

from tools import create_and_mix_color_scale_html

center_alingment = Alignment(horizontal="center", vertical="center")
right_alingment = Alignment(horizontal="right", vertical="center")
bold_font = Font(name='Calibri', bold=True)
red_font = Font(name='Calibri', color='FFFF0000')


def jsonify_data_table(data):
    data_table = pd.DataFrame.from_dict(data, orient='columns')
    cols = list(data_table.columns.values)
    data_table = data_table[[cols[-1]] + cols[:-1]]

    return data_table.to_json()


def jsonify_models(standard_models, std_x, std_y):
    if standard_models is None:
        return ''

    models = dict()
    models["standard_concentrations"] = std_x
    models['data'] = {}
    for std_name in standard_models.keys():
        model = standard_models[std_name]
        A, B, C, D = model["A"], model["B"], model["C"], model["D"]
        R2 = model["R^2"]
        models['data'][f'{std_name}'] = \
            {
                'A': A, 'B': B, 'C': C, 'D': D, "R2": R2,
                'Y': std_y[std_name]
            }
    return json.dumps(models)


def jsonify_results(data):
    return json.dumps(data)


def write_models_data(ws, data):
    free_col = ws.max_column + 1
    free_row = 17

    header_cell = ws.cell(column=free_col, row=free_row, value="Concentrations of standard:")
    header_cell.alignment = right_alingment
    header_cell.font = Font(bold=True)
    ws.column_dimensions[header_cell.column_letter].width = str(len("Concentrations of standard"))
    for col, xi in enumerate(data['standard_concentrations'], free_col + 1):
        ws.cell(column=col, row=free_row, value=xi)

    for i, model_name in enumerate(data['data'].keys()):
        start_row = free_row + (8 * i) + 3
        model = data['data'][model_name]

        model_name_cell = ws.cell(column=free_col, row=start_row, value=f'Model: {model_name}')
        ws.merge_cells(
            start_row=start_row,
            start_column=free_col,
            end_row=start_row,
            end_column=free_col + len(data['standard_concentrations'])
        )
        model_name_cell.font = bold_font
        model_name_cell.alignment = center_alingment

        y_name_cell = ws.cell(column=free_col, row=start_row + 1, value='Y:')
        y_name_cell.alignment = right_alingment
        y_name_cell.font = bold_font
        for j, yj in enumerate(model['Y'], 1):
            ws.cell(column=free_col + j, row=start_row + 1, value=yj)

        parameters_name_cell = ws.cell(column=free_col, row=start_row + 4, value='Parameters:')
        parameters_name_cell.alignment = right_alingment
        parameters_name_cell.font = bold_font

        for j, value in enumerate(['A', 'B', 'C', 'D'], 1):
            parameter_name_cell = ws.cell(column=free_col + j, row=start_row + 3, value=value)
            parameter_name_cell.alignment = right_alingment
            parameter_name_cell.font = bold_font

        for j, value in enumerate([model['A'], model['B'], model['C'], model['D']], 1):
            ws.cell(column=free_col + j, row=start_row + 4, value=value)

        r2_name_cell = ws.cell(column=free_col, row=start_row + 6, value='R2:')
        r2_name_cell.alignment = right_alingment
        r2_name_cell.font = bold_font
        ws.cell(column=free_col + 1, row=start_row + 6, value=model['R2'])


def write_results_data(ws, data):
    traces_pairs = list(zip_longest(*[iter(data.keys())] * 2, fillvalue=None))
    max_col = 1
    for pair in traces_pairs:
        start_row = max([cell.row for row in ws.iter_rows(min_col=max_col, max_col=max_col + 6, values_only=False)
                         for cell in row if cell.value is not None], default=0) + 2
        for i, trace in enumerate(pair):
            if trace is not None:
                x, y, labels = data[trace]['x'], data[trace]['y'], data[trace]['labels']
                start_col = max_col + (3 * i)
                trace_name_cell = ws.cell(column=start_col, row=start_row, value=trace)
                ws.merge_cells(
                    start_row=start_row,
                    start_column=start_col,
                    end_row=start_row,
                    end_column=start_col + 2
                )
                trace_name_cell.alignment = center_alingment
                trace_name_cell.font = bold_font
                for j, col_label in enumerate(['Label:', 'X:', 'Y:']):
                    cell = ws.cell(column=start_col + j, row=start_row + 1, value=col_label)
                    cell.alignment = center_alingment
                    cell.font = bold_font
                for j, xj, yj, labelj in zip(range(len(x)), x, y, labels):
                    ws.cell(column=start_col, row=start_row + 2 + j, value=labelj)
                    if xj is not None:
                        ws.cell(column=start_col + 1, row=start_row + 2 + j, value=xj)
                    ws.cell(column=start_col + 2, row=start_row + 2 + j, value=yj)
                tab = Table(displayName=f"{trace.replace(' ', '')}", ref="{}{}:{}{}".format(
                    get_column_letter(start_col),
                    start_row+1,
                    get_column_letter(start_col+2),
                    start_row+1+len(x)
                )
                            )
                ws.add_table(tab)

def write_std_curves_data(ws, model_data, results_data):
    trace_x = list(xi for trace in results_data.keys() for xi in results_data[trace]['x'] if xi is not None)
    max_value = max(
            *model_data['standard_concentrations'],
            *trace_x
        )
    x = np.arange(0, max_value+.01, .01).tolist()
    ws.cell(column=1, row=1, value='X')
    ws.cell(column=1, row=1, value='X')
    for i, xi in enumerate(x, 2):
        ws.cell(column=1, row=i, value=xi)

    for j, model_name in enumerate(model_data['data'].keys(), 2):
        ws.cell(column=j, row=1, value=model_name)
        model = model_data['data'][model_name]
        A, B, C, D = model['A'], model['B'], model['C'], model['D']
        for i, xi in enumerate(x, 2):
            val = ((A-D)/(1.0+((xi/C)**B))) + D
            ws.cell(column=j, row=i, value=val)

def plot_data(wb, anchor=None):
    std_ws = wb['STD curves Data']
    data_ws = wb['Data']
    x = std_ws['A'][1:]

    chart = ScatterChart()
    chart.display_blanks='gap'
    chart.style = 1
    chart.title = "Scatter Chart"
    chart.x_axis.title = 'x axis'
    chart.y_axis.title = 'y axis'
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = x[-1].value

    trace_colors = create_and_mix_color_scale_html(36)

    xvalues = Reference(std_ws, min_col=1, min_row=2, max_row=len(x)+1)
    for i in range(2, std_ws.max_column+1):
        color = trace_colors.pop()
        values = Reference(std_ws, min_col=i, min_row=1, max_row=len(x)+1)
        series = Series(values, xvalues, title_from_data=True)
        lineProp = drawing.line.LineProperties(solidFill=color, w=1)
        series.graphicalProperties.line = lineProp
        chart.series.append(series)

    for j, trace in enumerate(data_ws._tables):
        color = trace_colors.pop()
        start_ref, end_ref = trace.ref.split(':')
        start_coordinates, end_coordinates = coordinate_from_string(start_ref), coordinate_from_string(end_ref)
        start_row = start_coordinates[1]
        start_column = column_index_from_string(start_coordinates[0])

        title_cell = data_ws.cell(row=start_row-1, column=start_column)
        end_row = end_coordinates[1]
        end_column = column_index_from_string(end_coordinates[0])
        xvalues = Reference(data_ws, min_col=start_column+1, min_row=start_row+1, max_row=end_row)
        values = Reference(data_ws, min_col=end_column, min_row=start_row+1, max_row=end_row)
        series = Series(values, xvalues, title=title_cell.value)
        series.marker.symbol = "circle"
        series.marker.graphicalProperties.solidFill = color  # Marker filling
        series.marker.graphicalProperties.line.solidFill = color  # Marker outline
        series.graphicalProperties.line.noFill = True
        chart.series.append(series)

    data_ws.add_chart(chart, anchor)
